# -*- coding: utf-8 -*-
import io
import pandas as pd
import streamlit as st
from datetime import datetime
from openpyxl import load_workbook

st.set_page_config(page_title="储位利用率计算器", layout="wide")

st.title("📦 储位利用率计算器（支持 LAX1 / LAX2 / LAX4 / LAX5）")

with st.expander("使用说明", expanded=True):
    st.markdown("""
**步骤：**
1. 上传两份文件：**储位信息表** 与 **库存表**（Excel .xlsx）。  
2. 从下拉框选择 **筛选规则**（`LAX1` / `LAX2` / `LAX4` / `LAX5`）。  
3. 点击 **开始计算**，在页面查看结果并可下载 Excel。  

**字段要求：**
- 储位信息表（单位：毫米）：`储位编码`，`货架类型`，`长`，`宽`，`高`，`填充率` 
- 库存表（尺寸单位：英寸）：`储位编码`，`京东商品编码`，`长`，`宽`，`高`，`库存量`
""")

# ---------- 文件上传 ----------
col_u1, col_u2 = st.columns(2)
with col_u1:
    file_storage = st.file_uploader("上传储位信息表（.xlsx）", type=["xlsx"], key="storage")
with col_u2:
    file_inventory = st.file_uploader("上传库存表（.xlsx）", type=["xlsx"], key="inventory")

# ---------- 规则选择 ----------
rule = st.selectbox("选择筛选规则", options=["LAX1", "LAX2", "LAX4", "LAX5"], index=3)

# ---------- 工具函数 ----------
def ensure_columns(df, required_cols, name):
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        st.error(f"❌ {name} 缺少必需列：{missing}")
        st.stop()

def apply_rule(df, rule_name):
    """
    筛选规则：
    - LAX1：A 段在 A70 ~ A99（含）之间
    - LAX2：货架类型 in ["1窄巷道横梁式货架", "3搁板货架"]
    - LAX4：全部
    - LAX5：货架类型 in ["1单深横梁式货架"]
    """
    df = df.copy()

    if rule_name == "LAX1":
        # ✅ 修复：extract 返回 DataFrame 导致 mask 变二维的问题
        # 从“储位编码”中解析 A 段：如 A70-R01-L02-B03 → 70
        a_num = (
            df["储位编码"]
            .astype(str)
            .str.split("-", expand=True)[0]       # 取 A 段
            .str.extract(r"(\d+)")[0]             # ✅ 取第一列，变成 Series
        )
        a_num = pd.to_numeric(a_num, errors="coerce")
        mask = a_num.between(70, 99, inclusive="both").fillna(False)

    elif rule_name == "LAX2":
        mask = df["货架类型"].isin(["1窄巷道横梁式货架", "3搁板货架"])

    elif rule_name == "LAX4":
        mask = df.index == df.index  # 全部

    elif rule_name == "LAX5":
        mask = df["货架类型"].isin(["1单深横梁式货架"])

    else:
        mask = df.index == df.index

    return df.loc[mask].copy()

def compute_capacity(df_storage_filtered):
    # 储位体积（毫米->立方米）
    for c in ["长", "宽", "高", "填充率"]:
        df_storage_filtered[c] = pd.to_numeric(df_storage_filtered[c], errors="coerce")
    df_storage_filtered["储位体积"] = (
        df_storage_filtered["长"] * df_storage_filtered["宽"] * df_storage_filtered["高"]
    ) * df_storage_filtered["填充率"] / (1000 ** 3)
    df_storage_filtered.loc[df_storage_filtered["储位体积"] <= 0, "储位体积"] = pd.NA
    return df_storage_filtered

def sort_by_arlb(df):
    split_cols = df["储位编码"].astype(str).str.split("-", expand=True)
    if split_cols.shape[1] != 4:
        st.warning("⚠️ 储位编码未按 A-R-L-B 四段格式分列，将跳过排序。")
        df["_order"] = range(len(df))
        return df
    split_cols.columns = ["A", "R", "L", "B"]
    for col in ["A", "R", "L", "B"]:
        split_cols[col] = split_cols[col].str.extract(r"(\d+)").astype(float)
    df = pd.concat([df, split_cols], axis=1)
    df = df.sort_values(by=["A", "R", "L", "B"], ascending=True)
    df["_order"] = range(len(df))
    df = df.drop(columns=["A", "R", "L", "B"])
    return df

def compute_inventory_volume(df_inventory):
    for c in ["长", "宽", "高", "库存量"]:
        df_inventory[c] = pd.to_numeric(df_inventory[c], errors="coerce")
    df_inventory["库存体积"] = (
        df_inventory["长"] * df_inventory["宽"] * df_inventory["高"] * df_inventory["库存量"]
    ) * (0.0254 ** 3)
    return df_inventory

def aggregate_inventory_by_slot(df_inventory):
    return df_inventory.groupby("储位编码", as_index=False)["库存体积"].sum()

def compute_distinct_sku_count(df_inventory):
    sku_count = df_inventory.groupby("储位编码")["京东商品编码"].nunique().reset_index()
    sku_count.rename(columns={"京东商品编码": "储位SKU数量"}, inplace=True)
    return sku_count

def finalize_table(df_storage_sorted, inv_agg, sku_count):
    base = pd.merge(
        df_storage_sorted[["储位编码", "储位体积", "_order"]],
        inv_agg, on="储位编码", how="left"
    )
    base["储位利用率"] = (base["库存体积"] / base["储位体积"]) * 100
    base["储位利用率"] = base["储位利用率"].fillna(0).round(2)
    base = pd.merge(base, sku_count, on="储位编码", how="left")
    base["储位SKU数量"] = base["储位SKU数量"].fillna(0).astype(int)
    base = base.sort_values("_order").drop(columns=["_order"]).reset_index(drop=True)
    base = base[["储位编码", "储位体积", "库存体积", "储位利用率", "储位SKU数量"]]
    return base

def style_percent_to_str(df):
    df = df.copy()
    df["储位利用率"] = df["储位利用率"].map(lambda x: f"{x:.2f}%")
    return df

def to_excel_bytes_with_percent(df, sheet_name):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb[sheet_name]

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if "储位利用率" in header:
        col_idx = header.index("储位利用率") + 1
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, values_only=False):
            cell = row[0]
            val = cell.value
            try:
                num = float(val) / 100.0
            except Exception:
                num = 0.0
            cell.value = num
            cell.number_format = "0.00%"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ---------- 主逻辑 ----------
if st.button("▶️ 开始计算", type="primary"):
    if not file_storage or not file_inventory:
        st.error("请先上传 **储位信息表** 和 **库存表**。")
        st.stop()

    try:
        df_storage = pd.read_excel(file_storage)
        df_inventory = pd.read_excel(file_inventory)
    except Exception as e:
        st.error(f"读取 Excel 失败：{e}")
        st.stop()

    ensure_columns(df_storage, ["储位编码", "货架类型", "长", "宽", "高"], "储位信息表")
    ensure_columns(df_inventory, ["储位编码", "京东商品编码", "长", "宽", "高", "库存量"], "库存表")

    df_storage_filtered = apply_rule(df_storage, rule)
    st.success(f"已按规则 **{rule}** 筛选储位，匹配行数：{len(df_storage_filtered):,}")

    df_storage_cap = compute_capacity(df_storage_filtered)
    df_storage_sorted = sort_by_arlb(df_storage_cap)

    df_inventory_vol = compute_inventory_volume(df_inventory)
    inv_agg = aggregate_inventory_by_slot(df_inventory_vol)
    sku_count = compute_distinct_sku_count(df_inventory_vol)

    result = finalize_table(df_storage_sorted, inv_agg, sku_count)

    st.subheader("结果预览（前 100 行）")
    st.dataframe(style_percent_to_str(result.head(100)), use_container_width=True)

    today = datetime.now().strftime("%Y%m%d")
    sheet_name = f"{rule}_储位利用率"
    out_name = f"{rule}储位利用率表_{today}.xlsx"
    xlsx_bytes = to_excel_bytes_with_percent(result, sheet_name=sheet_name)
    st.download_button(
        "💾 下载Excel（百分比格式）",
        data=xlsx_bytes,
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    with st.expander("查看统计信息"):
        total_slots = len(result)
        filled_slots = (result["库存体积"].fillna(0) > 0).sum()
        over_100 = (result["储位利用率"].fillna(0) > 100).sum()
        st.write(f"- 储位总数：{total_slots:,}")
        st.write(f"- 有库存的储位：{filled_slots:,}")
        st.write(f"- 利用率 > 100% 的储位：{over_100:,}")
