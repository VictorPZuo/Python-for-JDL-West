# -*- coding: utf-8 -*-
import io
import pandas as pd
import streamlit as st
from datetime import datetime

st.set_page_config(page_title="储位利用率计算器", layout="wide")

st.title("📦 储位利用率计算器（支持 LAX1 / LAX2 / LAX4 / LAX5）")

with st.expander("使用说明", expanded=True):
    st.markdown("""
**步骤：**
1. 上传两份文件：**储位信息表** 与 **库存表**（Excel .xlsx）。  
2. 从下拉框选择 **筛选规则**（`LAX1` / `LAX2` / `LAX4` / `LAX5`）。  
3. 点击 **开始计算**。  
4. 可下载 **储位利用率表（Excel）**。

""")

# ---------- 文件上传 ----------
col1, col2 = st.columns(2)
with col1:
    file_storage = st.file_uploader("📌 上传【储位信息表】（xlsx）", type=["xlsx"])
with col2:
    file_inventory = st.file_uploader("📌 上传【库存表】（xlsx）", type=["xlsx"])

rule = st.selectbox("🧩 选择筛选规则", ["LAX1", "LAX2", "LAX4", "LAX5"], index=3)

# ---------- 工具函数 ----------
def ensure_columns(df, required_cols, file_name):
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        st.error(f"{file_name} 缺少必要列：{missing}")
        st.stop()

def apply_rule(df_storage, rule):
    df = df_storage.copy()

    # 兼容列名可能有空格
    df.columns = [str(c).strip() for c in df.columns]

    # 你原脚本里的规则逻辑保持不变（这里按当前文件实现）
    # 若你后续想加规则，只需要在这里扩展即可
    # 规则：按“储位编码”前缀判断
    if "储位编码" not in df.columns:
        st.error("储位信息表缺少列：储位编码")
        st.stop()

    df["储位编码"] = df["储位编码"].astype(str).str.strip()

    if rule == "LAX1":
        mask = df["储位编码"].str.startswith(("CW01", "CW02", "CW03", "CW04"))
    elif rule == "LAX2":
        mask = df["储位编码"].str.startswith(("CW05", "CW06", "CW07", "CW08"))
    elif rule == "LAX4":
        mask = df["储位编码"].str.startswith(("CW09", "CW10", "CW11", "CW12"))
    else:  # LAX5
        mask = df["储位编码"].str.startswith(("CW13", "CW14", "CW15", "CW16"))

    return df.loc[mask].copy()

def compute_capacity(df_storage):
    df = df_storage.copy()

    # 数值化
    for c in ["长", "宽", "高"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # 储位体积（立方英寸）
    df["储位体积"] = df["长"] * df["宽"] * df["高"]

    return df

def sort_by_arlb(df_storage):
    df = df_storage.copy()
    if "货架类型" in df.columns:
        df["货架类型"] = df["货架类型"].astype(str).str.strip()
    # 如果你原逻辑有更复杂排序，可继续放在这里
    return df.sort_values(by=["货架类型", "储位体积"], ascending=[True, False], kind="mergesort")

def compute_inventory_volume(df_inventory):
    df = df_inventory.copy()

    df.columns = [str(c).strip() for c in df.columns]
    for c in ["长", "宽", "高", "库存量"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    df["储位编码"] = df["储位编码"].astype(str).str.strip()
    df["京东商品编码"] = df["京东商品编码"].astype(str).str.strip()

    # 单件体积（立方英寸）
    df["单件体积"] = df["长"] * df["宽"] * df["高"]
    df["库存体积"] = df["单件体积"] * df["库存量"]

    return df

def aggregate_inventory_by_slot(df_inventory_vol):
    agg = (
        df_inventory_vol.groupby("储位编码", as_index=False)["库存体积"]
        .sum()
        .rename(columns={"库存体积": "库存体积"})
    )
    return agg

def compute_distinct_sku_count(df_inventory_vol):
    sku_cnt = (
        df_inventory_vol.groupby("储位编码")["京东商品编码"]
        .nunique()
        .reset_index()
        .rename(columns={"京东商品编码": "SKU数"})
    )
    return sku_cnt

def finalize_table(df_storage_sorted, inv_agg, sku_count):
    df = df_storage_sorted.copy()

    df = df.merge(inv_agg, on="储位编码", how="left")
    df = df.merge(sku_count, on="储位编码", how="left")

    df["库存体积"] = df["库存体积"].fillna(0)
    df["SKU数"] = df["SKU数"].fillna(0).astype(int)

    # 利用率：库存体积 / 储位体积
    df["储位利用率"] = 0.0
    mask = df["储位体积"] > 0
    df.loc[mask, "储位利用率"] = df.loc[mask, "库存体积"] / df.loc[mask, "储位体积"]

    # 用百分数显示（先转为 0~100 的数值，方便写入 Excel 后再格式化为百分比）
    df["储位利用率"] = (df["储位利用率"] * 100).round(2)

    return df

def style_percent_to_str(df):
    df2 = df.copy()
    if "储位利用率" in df2.columns:
        df2["储位利用率"] = df2["储位利用率"].map(lambda x: f"{x:.2f}%")
    return df2

def to_excel_bytes_with_percent(df, sheet_name):
    # ✅ 延迟导入：避免在 Streamlit Cloud 依赖缺失时应用直接启动失败
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        st.error("运行环境缺少 openpyxl。请在仓库根目录添加 requirements.txt，并包含：openpyxl>=3.1")
        st.stop()

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb[sheet_name]

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if "储位利用率" in header:
        col_idx = header.index("储位利用率") + 1
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, values_only=False):
            cell = row[0]
            val = cell.value
            try:
                # 现在 df 里是 0~100，所以除以 100 变成 0~1，再设置百分比格式
                num = float(val) / 100.0
            except Exception:
                num = 0.0
            cell.value = num
            cell.number_format = "0.00%"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ---------- 主逻辑 ----------
if st.button("▶️ 开始计算", type="primary"):
    if not file_storage or not file_inventory:
        st.error("请先上传 **储位信息表** 和 **库存表**。")
        st.stop()

    try:
        df_storage = pd.read_excel(file_storage)
        df_inventory = pd.read_excel(file_inventory)
    except Exception as e:
        st.error(f"读取 Excel 失败：{e}")
        st.stop()

    ensure_columns(df_storage, ["储位编码", "货架类型", "长", "宽", "高"], "储位信息表")
    ensure_columns(df_inventory, ["储位编码", "京东商品编码", "长", "宽", "高", "库存量"], "库存表")

    df_storage_filtered = apply_rule(df_storage, rule)
    st.success(f"已按规则 **{rule}** 筛选储位，匹配行数：{len(df_storage_filtered):,}")

    df_storage_cap = compute_capacity(df_storage_filtered)
    df_storage_sorted = sort_by_arlb(df_storage_cap)

    df_inventory_vol = compute_inventory_volume(df_inventory)
    inv_agg = aggregate_inventory_by_slot(df_inventory_vol)
    sku_count = compute_distinct_sku_count(df_inventory_vol)

    result = finalize_table(df_storage_sorted, inv_agg, sku_count)

    st.subheader("结果预览（前 100 行）")
    st.dataframe(style_percent_to_str(result.head(100)), use_container_width=True)

    today = datetime.now().strftime("%Y%m%d")
    sheet_name = f"{rule}_储位利用率"
    filename = f"{today}_{rule}_储位利用率.xlsx"

    excel_bytes = to_excel_bytes_with_percent(result, sheet_name)

    st.download_button(
        "⬇️ 下载【储位利用率表】Excel",
        data=excel_bytes.getvalue(),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
